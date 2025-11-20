#!/usr/bin/env python3
"""
Phase 3: Data Merge Script

This script merges BACB supervision data with transformed supervision data,
creates the final output file with all columns including BACB data.

Usage:
    python merge_data.py [--transformed-input PATH] [--bacb-input PATH] [--output PATH]
"""

import pandas as pd
import os
import shutil
import logging
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill


def setup_logging(log_dir: str = None) -> logging.Logger:
    """Set up logging configuration."""
    # Use root logs directory if not specified
    if log_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Go up from scripts_notebooks/prod to project root
        project_root = os.path.dirname(os.path.dirname(script_dir))
        log_dir = os.path.join(project_root, 'logs')
    
    # Ensure logs directory exists
    os.makedirs(log_dir, exist_ok=True)
    
    # Create log file path
    log_file = os.path.join(log_dir, 'merge_data.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


def adjust_column_widths(ws, logger):
    """
    Adjust column widths to fit the content (header + data).
    
    Args:
        ws: openpyxl worksheet object
        logger: Logger instance
    """
    from openpyxl.utils import get_column_letter
    
    # Iterate through all columns
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check header cell
        header_cell = ws[f'{column_letter}1']
        if header_cell.value:
            max_length = max(max_length, len(str(header_cell.value)))
        
        # Check data cells (sample first 100 rows for performance)
        for row_idx, cell in enumerate(column[1:], start=2):  # Skip header row
            if cell.value is not None:
                cell_value = str(cell.value)
                # For numbers, consider formatted length
                if isinstance(cell.value, (int, float)):
                    # Estimate formatted length (add some padding for decimals)
                    cell_length = len(f"{cell.value:.2f}")
                else:
                    cell_length = len(cell_value)
                max_length = max(max_length, cell_length)
            
            # Sample only first 100 data rows for performance
            if row_idx > 100:
                break
        
        # Set column width (add padding, min 10, max 50)
        width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = width


def save_to_google_drive_folder(source_file: str, target_folder: str, logger: logging.Logger):
    """
    Save the Excel file to Google Drive folder and archive existing files.
    
    Args:
        source_file (str): Path to the source Excel file
        target_folder (str): Path to the target Google Drive folder
        logger: Logger instance
    """
    # Ensure target folder exists
    os.makedirs(target_folder, exist_ok=True)
    
    # Create archive folder in target location
    archive_folder = os.path.join(target_folder, 'archived')
    os.makedirs(archive_folder, exist_ok=True)
    
    # Archive existing .xlsx files in target folder
    output_filename = os.path.basename(source_file)
    if os.path.exists(target_folder):
        existing_files = [f for f in os.listdir(target_folder) 
                        if f.endswith('.xlsx') and f != output_filename]
        
        for file in existing_files:
            source_path = os.path.join(target_folder, file)
            archive_path = os.path.join(archive_folder, file)
            
            # If file already exists in archive, add timestamp to avoid conflicts
            if os.path.exists(archive_path):
                name, ext = os.path.splitext(file)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                archive_path = os.path.join(archive_folder, f'{name}_{timestamp}{ext}')
            
            shutil.move(source_path, archive_path)
            logger.info(f"Archived existing file in Google Drive folder: {file}")
    
    # Copy the file to Google Drive folder
    target_file = os.path.join(target_folder, output_filename)
    shutil.copy2(source_file, target_file)
    logger.info(f"Saved file to Google Drive folder: {target_file}")


def merge_data(transformed_df: pd.DataFrame, bacb_df: pd.DataFrame, logger: logging.Logger) -> pd.DataFrame:
    """
    Merge BACB supervision data with transformed supervision data.
    
    Joins BACB data onto the transformed DataFrame using DirectProviderId.
    Handles missing values for providers without BACB supervision.
    
    Args:
        transformed_df (pd.DataFrame): Transformed supervision data
        bacb_df (pd.DataFrame): BACB supervision data with ProviderContactId
        logger: Logger instance
        
    Returns:
        pd.DataFrame: Final merged DataFrame with all columns including BACB data
    """
    logger.info("="*50)
    logger.info("Phase 3: Data Merge")
    logger.info("="*50)
    
    # Join BACB data to transformed DataFrame
    logger.info("Merging BACB supervision data...")
    merged_df = transformed_df.merge(
        bacb_df,
        left_on='DirectProviderId',
        right_on='ProviderContactId',
        how='left'
    )
    
    # Fill NaN values for BACB columns (providers without BACB supervision)
    merged_df['BACBSupervisionCodes_binary'] = merged_df['BACBSupervisionCodes_binary'].fillna(0).astype(int)
    merged_df['BACBSupervisionHours'] = merged_df['BACBSupervisionHours'].fillna(0.0)
    
    # Rename column and convert 1/0 to Yes/No
    merged_df['BACBSupervisionCodesOccurred'] = merged_df['BACBSupervisionCodes_binary'].map({1: 'Yes', 0: 'No'})
    merged_df.drop(columns=['BACBSupervisionCodes_binary'], inplace=True, errors='ignore')
    
    # Drop the ProviderContactId column since we're using DirectProviderId
    merged_df.drop(columns=['ProviderContactId'], inplace=True, errors='ignore')
    
    # Add TotalSupervisionHours column (sum of SupervisionHours and BACBSupervisionHours)
    if 'SupervisionHours' in merged_df.columns and 'BACBSupervisionHours' in merged_df.columns:
        merged_df['TotalSupervisionHours'] = merged_df['SupervisionHours'].fillna(0) + merged_df['BACBSupervisionHours'].fillna(0)
        logger.info("Added TotalSupervisionHours column (SupervisionHours + BACBSupervisionHours)")
    
    # Calculate percentage of direct hours supervised using TotalSupervisionHours
    if 'TotalSupervisionHours' in merged_df.columns and 'DirectHours' in merged_df.columns:
        merged_df['TotalSupervisionPercent'] = round(
            100 * (merged_df['TotalSupervisionHours'] / merged_df['DirectHours'].replace(0, pd.NA)), 2
        )
        logger.info("Added TotalSupervisionPercent column (100 * TotalSupervisionHours / DirectHours)")
    
    # Reorder columns to include BACB data - TotalSupervisionPercent should be last
    column_order = [
        'Clinic', 'DirectProviderId', 'DirectProviderName', 
        'DirectHours', 'SupervisionHours',
        'BACBSupervisionCodesOccurred', 'BACBSupervisionHours', 'TotalSupervisionHours'
    ]
    # Only include columns that exist in the order list
    column_order = [col for col in column_order if col in merged_df.columns]
    # Add any remaining columns (except TotalSupervisionPercent which goes last)
    remaining_cols = [col for col in merged_df.columns if col not in column_order and col != 'TotalSupervisionPercent']
    column_order.extend(remaining_cols)
    # Ensure TotalSupervisionPercent is always last if it exists
    if 'TotalSupervisionPercent' in merged_df.columns:
        column_order.append('TotalSupervisionPercent')
    merged_df = merged_df[column_order]
    
    logger.info(f"Data merge completed. Final dataset has {len(merged_df)} rows with BACB columns.")
    return merged_df


def merge_data_main(transformed_df: pd.DataFrame = None, bacb_df: pd.DataFrame = None,
                   transformed_file: str = None, bacb_file: str = None, save_file: bool = True) -> pd.DataFrame:
    """
    Main function to merge data.
    
    Args:
        transformed_df (pd.DataFrame, optional): Transformed DataFrame. If None, will read from transformed_file.
        bacb_df (pd.DataFrame, optional): BACB DataFrame. If None, will read from bacb_file.
        transformed_file (str, optional): Transformed CSV file path. Used if transformed_df is None.
        bacb_file (str, optional): BACB CSV file path. Used if bacb_df is None.
        save_file (bool): Whether to save file to disk. Default True.
        
    Returns:
        pd.DataFrame: Final merged DataFrame
    """
    # Set up logging
    logger = setup_logging()
    
    # Get transformed data
    if transformed_df is None:
        if transformed_file is None:
            today = datetime.now().strftime('%Y-%m-%d')
            # Try .xlsx first, fallback to .csv for backward compatibility
            xlsx_file = f'../../data/transformed_supervision_daily/daily_supervision_hours_transformed_{today}.xlsx'
            csv_file = f'../../data/transformed_supervision_daily/daily_supervision_hours_transformed_{today}.csv'
            transformed_file = xlsx_file if os.path.exists(xlsx_file) else csv_file
        
        if not os.path.exists(transformed_file):
            logger.error(f"Transformed input file not found: {transformed_file}")
            raise FileNotFoundError(f"Transformed input file not found: {transformed_file}")
        
        logger.info(f"Reading transformed data from: {transformed_file}")
        # Read CSV or Excel based on file extension
        if transformed_file.endswith('.xlsx'):
            transformed_df = pd.read_excel(transformed_file, engine='openpyxl')
        else:
            transformed_df = pd.read_csv(transformed_file)
        logger.info(f"Loaded {len(transformed_df)} rows from transformed file")
    else:
        logger.info(f"Using provided transformed DataFrame with {len(transformed_df)} rows")
    
    # Get BACB data
    if bacb_df is None:
        if bacb_file is None:
            today = datetime.now().strftime('%Y-%m-%d')
            bacb_file = f'../../data/raw_pulls/bacb_supervision_hours_{today}.csv'
        
        if not os.path.exists(bacb_file):
            logger.error(f"BACB input file not found: {bacb_file}")
            raise FileNotFoundError(f"BACB input file not found: {bacb_file}")
        
        logger.info(f"Reading BACB data from: {bacb_file}")
        bacb_df = pd.read_csv(bacb_file)
        logger.info(f"Loaded {len(bacb_df)} rows from BACB file")
    else:
        logger.info(f"Using provided BACB DataFrame with {len(bacb_df)} rows")
    
    # Merge data
    final_df = merge_data(transformed_df, bacb_df, logger)
    
    if save_file:
        # Archive existing files before saving new one
        today = datetime.now().strftime('%Y-%m-%d')
        output_file = f'../../data/transformed_supervision_daily/daily_supervision_hours_transformed_{today}.xlsx'
        archive_folder = f'../../data/transformed_supervision_daily/archived'
        
        # Ensure directories exist
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        os.makedirs(archive_folder, exist_ok=True)
        
        # Archive existing files (CSV and XLSX, excluding the one we're about to create)
        output_filename = os.path.basename(output_file)
        if os.path.exists(os.path.dirname(output_file)):
            existing_files = [f for f in os.listdir(os.path.dirname(output_file)) 
                            if (f.endswith('.csv') or f.endswith('.xlsx')) and f != output_filename]
            
            for file in existing_files:
                source_path = os.path.join(os.path.dirname(output_file), file)
                archive_path = os.path.join(archive_folder, file)
                
                # If file already exists in archive, add timestamp to avoid conflicts
                if os.path.exists(archive_path):
                    name, ext = os.path.splitext(file)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    archive_path = os.path.join(archive_folder, f'{name}_{timestamp}{ext}')
                
                shutil.move(source_path, archive_path)
                logger.info(f"Archived existing file: {file}")
        
        # Group data by Clinic and save as Excel with separate sheets
        if 'Clinic' in final_df.columns:
            # Get unique clinics that actually have data
            # Filter to only clinics that have at least one row
            clinics_with_data = final_df.groupby('Clinic').size()
            clinics = clinics_with_data[clinics_with_data > 0].index.tolist()
            logger.info(f"Saving Excel file with {len(clinics)} clinic sheets (clinics with data)")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for clinic in sorted(clinics):
                    clinic_data = final_df[final_df['Clinic'] == clinic].copy()
                    
                    # This should always have data since we filtered above, but double-check
                    if len(clinic_data) == 0:
                        logger.warning(f"  - Skipping sheet for '{clinic}' (no data found)")
                        continue
                    
                    # Sort by TotalSupervisionPercent (lowest values first)
                    if 'TotalSupervisionPercent' in clinic_data.columns:
                        clinic_data = clinic_data.sort_values('TotalSupervisionPercent', ascending=True, na_position='last')
                        logger.info(f"  - Sorted {len(clinic_data)} rows by TotalSupervisionPercent (ascending)")
                    
                    # Excel sheet names must be <= 31 characters and can't contain certain characters
                    # Clean the clinic name for the sheet name
                    sheet_name = str(clinic)[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                    clinic_data.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"  - Saved {len(clinic_data)} rows to sheet '{sheet_name}'")
            
            # Add conditional formatting after writing
            logger.info("Adding conditional formatting to Excel file...")
            wb = load_workbook(output_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Find the column index for TotalSupervisionPercent
                header_row = 1
                pct_col_idx = None
                for cell in ws[header_row]:
                    if cell.value == 'TotalSupervisionPercent':
                        pct_col_idx = cell.column_letter
                        break
                
                if pct_col_idx:
                    # Get the last row with data
                    max_row = ws.max_row
                    # Skip conditional formatting if there are no data rows (only header)
                    if max_row <= 1:
                        logger.info(f"  - Skipping conditional formatting for sheet '{sheet_name}' (no data rows)")
                    else:
                        data_range = f'{pct_col_idx}2:{pct_col_idx}{max_row}'
                        
                        # Add conditional formatting with discrete ranges:
                        # 0-5% = Red background (inclusive)
                        # >5% and <10% = Yellow background
                        # >=10% = Green background
                        
                        # Red background for <= 5%
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        red_rule = CellIsRule(operator='lessThanOrEqual', formula=[5.0], fill=red_fill)
                        ws.conditional_formatting.add(data_range, red_rule)
                        
                        # Yellow background for > 5% and < 10% (using formula for proper AND logic)
                        yellow_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                        # Use FormulaRule with relative reference - Excel will adjust for each cell
                        yellow_formula = f'AND({pct_col_idx}2>5, {pct_col_idx}2<10)'
                        yellow_rule = FormulaRule(formula=[yellow_formula], fill=yellow_fill)
                        ws.conditional_formatting.add(data_range, yellow_rule)
                        
                        # Green background for >= 10%
                        green_fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                        green_rule = CellIsRule(operator='greaterThanOrEqual', formula=[10.0], fill=green_fill)
                        ws.conditional_formatting.add(data_range, green_rule)
                        
                        logger.info(f"  - Added conditional formatting to column {pct_col_idx} in sheet '{sheet_name}' (0-5% red, >5-<10% yellow, >=10% green)")
                
                # Find the column index for BACBSupervisionCodesOccurred and add conditional formatting
                bacb_col_idx = None
                for cell in ws[header_row]:
                    if cell.value == 'BACBSupervisionCodesOccurred':
                        bacb_col_idx = cell.column_letter
                        break
                
                if bacb_col_idx:
                    max_row = ws.max_row
                    # Skip conditional formatting if there are no data rows (only header)
                    if max_row <= 1:
                        logger.info(f"  - Skipping BACB conditional formatting for sheet '{sheet_name}' (no data rows)")
                    else:
                        bacb_data_range = f'{bacb_col_idx}2:{bacb_col_idx}{max_row}'
                        
                        # Red background for "No"
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        red_formula = f'{bacb_col_idx}2="No"'
                        red_rule = FormulaRule(formula=[red_formula], fill=red_fill)
                        ws.conditional_formatting.add(bacb_data_range, red_rule)
                        
                        # Green background for "Yes"
                        green_fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                        green_formula = f'{bacb_col_idx}2="Yes"'
                        green_rule = FormulaRule(formula=[green_formula], fill=green_fill)
                        ws.conditional_formatting.add(bacb_data_range, green_rule)
                        
                        logger.info(f"  - Added conditional formatting to column {bacb_col_idx} in sheet '{sheet_name}' (No=red, Yes=green)")
                
                # Adjust column widths to fit content
                logger.info(f"  - Adjusting column widths for sheet '{sheet_name}'...")
                adjust_column_widths(ws, logger)
            
            wb.save(output_file)
            logger.info(f"Saved final merged data to Excel file: {output_file}")
            
            # Also save to Google Drive folder
            google_drive_folder = '/Users/davidjcox/Library/CloudStorage/GoogleDrive-dcox@mosaictherapy.com/.shortcut-targets-by-id/1Mh9gqV27KkEEuyX6M35_SB_vTErRz7Gm/DailyRBTTracking'
            try:
                save_to_google_drive_folder(output_file, google_drive_folder, logger)
            except Exception as e:
                logger.warning(f"Failed to save to Google Drive folder: {e}")
        else:
            # Fallback: save as single sheet if Clinic column doesn't exist
            logger.warning("'Clinic' column not found, saving as single sheet")
            
            # Sort by TotalSupervisionPercent (lowest values first)
            if 'TotalSupervisionPercent' in final_df.columns:
                final_df = final_df.sort_values('TotalSupervisionPercent', ascending=True, na_position='last')
                logger.info(f"Sorted {len(final_df)} rows by TotalSupervisionPercent (ascending)")
            
            final_df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Add conditional formatting
            logger.info("Adding conditional formatting to Excel file...")
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Find the column index for TotalSupervisionPercent
            header_row = 1
            pct_col_idx = None
            for cell in ws[header_row]:
                if cell.value == 'TotalSupervisionPercent':
                    pct_col_idx = cell.column_letter
                    break
            
            if pct_col_idx:
                max_row = ws.max_row
                # Skip conditional formatting if there are no data rows (only header)
                if max_row <= 1:
                    logger.info("Skipping conditional formatting (no data rows)")
                else:
                    data_range = f'{pct_col_idx}2:{pct_col_idx}{max_row}'
                    
                    # Add conditional formatting with discrete ranges:
                    # 0-5% = Red background (inclusive)
                    # >5% and <10% = Yellow background
                    # >=10% = Green background
                    
                    # Red background for <= 5%
                    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    red_rule = CellIsRule(operator='lessThanOrEqual', formula=[5.0], fill=red_fill)
                    ws.conditional_formatting.add(data_range, red_rule)
                    
                    # Yellow background for > 5% and < 10% (using formula for proper AND logic)
                    yellow_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                    # Use FormulaRule with relative reference - Excel will adjust for each cell
                    yellow_formula = f'AND({pct_col_idx}2>5, {pct_col_idx}2<10)'
                    yellow_rule = FormulaRule(formula=[yellow_formula], fill=yellow_fill)
                    ws.conditional_formatting.add(data_range, yellow_rule)
                    
                    # Green background for >= 10%
                    green_fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                    green_rule = CellIsRule(operator='greaterThanOrEqual', formula=[10.0], fill=green_fill)
                    ws.conditional_formatting.add(data_range, green_rule)
                    
                    logger.info(f"Added conditional formatting to column {pct_col_idx} (0-5% red, >5-<10% yellow, >=10% green)")
            
            # Find the column index for BACBSupervisionCodesOccurred and add conditional formatting
            bacb_col_idx = None
            for cell in ws[header_row]:
                if cell.value == 'BACBSupervisionCodesOccurred':
                    bacb_col_idx = cell.column_letter
                    break
            
            if bacb_col_idx:
                max_row = ws.max_row
                # Skip conditional formatting if there are no data rows (only header)
                if max_row <= 1:
                    logger.info("Skipping BACB conditional formatting (no data rows)")
                else:
                    bacb_data_range = f'{bacb_col_idx}2:{bacb_col_idx}{max_row}'
                    
                    # Red background for "No"
                    red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    red_formula = f'{bacb_col_idx}2="No"'
                    red_rule = FormulaRule(formula=[red_formula], fill=red_fill)
                    ws.conditional_formatting.add(bacb_data_range, red_rule)
                    
                    # Green background for "Yes"
                    green_fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                    green_formula = f'{bacb_col_idx}2="Yes"'
                    green_rule = FormulaRule(formula=[green_formula], fill=green_fill)
                    ws.conditional_formatting.add(bacb_data_range, green_rule)
                    
                    logger.info(f"Added conditional formatting to column {bacb_col_idx} (No=red, Yes=green)")
            
            # Adjust column widths to fit content
            logger.info("Adjusting column widths...")
            adjust_column_widths(ws, logger)
            
            wb.save(output_file)
            logger.info(f"Saved final merged data to: {output_file}")
            
            # Also save to Google Drive folder
            google_drive_folder = '/Users/davidjcox/Library/CloudStorage/GoogleDrive-dcox@mosaictherapy.com/.shortcut-targets-by-id/1Mh9gqV27KkEEuyX6M35_SB_vTErRz7Gm/DailyRBTTracking'
            try:
                save_to_google_drive_folder(output_file, google_drive_folder, logger)
            except Exception as e:
                logger.warning(f"Failed to save to Google Drive folder: {e}")
    
    logger.info("="*50)
    logger.info("Data merge completed successfully!")
    logger.info("="*50)
    
    return final_df


def main():
    """CLI entry point for merge_data.py"""
    parser = argparse.ArgumentParser(description='Merge transformed and BACB supervision data')
    parser.add_argument('--transformed-input', type=str,
                       default='../../data/transformed_supervision_daily/daily_supervision_hours_transformed_{date}.csv',
                       help='Input CSV file path for transformed data (use {date} placeholder)')
    parser.add_argument('--bacb-input', type=str,
                       default='../../data/raw_pulls/bacb_supervision_hours_{date}.csv',
                       help='Input CSV file path for BACB data (use {date} placeholder)')
    parser.add_argument('--output', type=str,
                       default='../../data/transformed_supervision_daily/daily_supervision_hours_transformed_{date}.xlsx',
                       help='Output Excel file path (use {date} placeholder)')
    
    args = parser.parse_args()
    
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        transformed_file = args.transformed_input.format(date=today)
        bacb_file = args.bacb_input.format(date=today)
        merge_data_main(transformed_file=transformed_file, bacb_file=bacb_file, save_file=True)
        return 0
    except Exception as e:
        logging.error(f"Error in data merge: {e}")
        raise


if __name__ == "__main__":
    exit(main())